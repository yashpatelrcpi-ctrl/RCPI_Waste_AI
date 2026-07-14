"""
Report Generation Module
Generate PDF and Excel reports for complaints, waste, and analytics
"""

import sqlite3
from datetime import datetime, timedelta
from typing import Optional, List
import json
from database import get_connection

class ReportGenerator:
    """Generate reports in various formats"""
    
    @staticmethod
    def get_complaint_summary(start_date: Optional[str] = None, 
                             end_date: Optional[str] = None) -> dict:
        """Get complaint summary data"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Default to last 30 days
            if not end_date:
                end_date = datetime.now()
            else:
                end_date = datetime.fromisoformat(end_date)
            
            if not start_date:
                start_date = end_date - timedelta(days=30)
            else:
                start_date = datetime.fromisoformat(start_date)
            # Use ISO strings for SQLite
            start_ts = start_date.isoformat()
            end_ts = end_date.isoformat()

            cursor.execute("PRAGMA table_info(complaints)")
            complaint_columns = {row[1] for row in cursor.fetchall()}
            use_date_filter = 'created_date' in complaint_columns

            # Total complaints (case-insensitive status checks)
            if use_date_filter:
                cursor.execute('''
                    SELECT COUNT(*), 
                           SUM(CASE WHEN LOWER(status)='pending' THEN 1 ELSE 0 END),
                           SUM(CASE WHEN LOWER(status)='resolved' THEN 1 ELSE 0 END),
                           SUM(CASE WHEN LOWER(status)='in_progress' THEN 1 ELSE 0 END)
                    FROM complaints 
                    WHERE created_date BETWEEN ? AND ?
                ''', (start_ts, end_ts))
            else:
                cursor.execute('''
                    SELECT COUNT(*), 
                           SUM(CASE WHEN LOWER(status)='pending' THEN 1 ELSE 0 END),
                           SUM(CASE WHEN LOWER(status)='resolved' THEN 1 ELSE 0 END),
                           SUM(CASE WHEN LOWER(status)='in_progress' THEN 1 ELSE 0 END)
                    FROM complaints
                ''')
            
            result = cursor.fetchone()
            total, pending, resolved, in_progress = result if result else (0, 0, 0, 0)
            
            # Complaints by type
            if use_date_filter:
                cursor.execute('''
                    SELECT waste_type, COUNT(*) 
                    FROM complaints 
                    WHERE created_date BETWEEN ? AND ?
                    GROUP BY waste_type
                ''', (start_ts, end_ts))
            else:
                cursor.execute('''
                    SELECT waste_type, COUNT(*) 
                    FROM complaints 
                    GROUP BY waste_type
                ''')
            
            by_type = {row[0]: row[1] for row in cursor.fetchall()}
            
            # Complaints by ward
            if use_date_filter:
                cursor.execute('''
                    SELECT ward, COUNT(*) 
                    FROM complaints 
                    WHERE created_date BETWEEN ? AND ?
                    GROUP BY ward
                    ORDER BY COUNT(*) DESC
                ''', (start_ts, end_ts))
            else:
                cursor.execute('''
                    SELECT ward, COUNT(*) 
                    FROM complaints 
                    GROUP BY ward
                    ORDER BY COUNT(*) DESC
                ''')
            
            by_ward = {row[0]: row[1] for row in cursor.fetchall()}
            
            # Average resolution time (may not exist on older schema)
            try:
                if use_date_filter:
                    cursor.execute('''
                        SELECT AVG(CAST((julianday(resolved_date) - julianday(created_date)) AS REAL))
                        FROM complaints
                        WHERE LOWER(status)='resolved' AND created_date BETWEEN ? AND ?
                    ''', (start_ts, end_ts))
                else:
                    cursor.execute('''
                        SELECT AVG(CAST((julianday(resolved_date) - julianday(created_date)) AS REAL))
                        FROM complaints
                        WHERE LOWER(status)='resolved'
                    ''')
                avg_resolution = cursor.fetchone()[0] or 0
                # avg_resolution is in days already via julianday difference
            except Exception:
                avg_resolution = 0
            
            conn.close()
            
            return {
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'summary': {
                    'total_complaints': total or 0,
                    'pending': pending or 0,
                    'resolved': resolved or 0,
                    'in_progress': in_progress or 0,
                    'resolution_rate': ((resolved or 0) / (total or 1) * 100) if total else 0
                },
                'by_type': by_type,
                'by_ward': by_ward,
                'avg_resolution_days': round(avg_resolution, 2)
            }
        
        except Exception as e:
            return {'error': str(e)}
    
    @staticmethod
    def get_waste_summary(start_date: Optional[str] = None,
                         end_date: Optional[str] = None) -> dict:
        """Get waste collection summary"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            if not end_date:
                end_date = datetime.now()
            else:
                end_date = datetime.fromisoformat(end_date)
            
            if not start_date:
                start_date = end_date - timedelta(days=30)
            else:
                start_date = datetime.fromisoformat(start_date)
            
            start_ts = start_date.isoformat()
            end_ts = end_date.isoformat()
            
            cursor.execute("PRAGMA table_info(waste_collection)")
            waste_columns = {row[1] for row in cursor.fetchall()}
            use_waste_date_filter = 'collection_date' in waste_columns

            # Total waste collected
            if use_waste_date_filter:
                cursor.execute('SELECT SUM(waste_quantity) FROM waste_collection WHERE collection_date BETWEEN ? AND ?', (start_ts, end_ts))
            else:
                cursor.execute('SELECT SUM(waste_quantity) FROM waste_collection')

            total_waste = cursor.fetchone()[0] or 0

            # Waste by type
            if use_waste_date_filter:
                cursor.execute('''
                    SELECT waste_type, SUM(waste_quantity), COUNT(*)
                    FROM waste_collection
                    WHERE collection_date BETWEEN ? AND ?
                    GROUP BY waste_type
                ''', (start_ts, end_ts))
            else:
                cursor.execute('''
                    SELECT waste_type, SUM(waste_quantity), COUNT(*)
                    FROM waste_collection
                    GROUP BY waste_type
                ''')

            by_type = {row[0]: {'quantity': row[1] or 0, 'collections': row[2] or 0}
                      for row in cursor.fetchall()}

            # Waste by ward
            if use_waste_date_filter:
                cursor.execute('''
                    SELECT ward, SUM(waste_quantity), COUNT(*)
                    FROM waste_collection
                    WHERE collection_date BETWEEN ? AND ?
                    GROUP BY ward
                    ORDER BY SUM(waste_quantity) DESC
                ''', (start_ts, end_ts))
            else:
                cursor.execute('''
                    SELECT ward, SUM(waste_quantity), COUNT(*)
                    FROM waste_collection
                    GROUP BY ward
                    ORDER BY SUM(waste_quantity) DESC
                ''')

            by_ward = {row[0]: {'quantity': row[1] or 0, 'collections': row[2] or 0}
                      for row in cursor.fetchall()}
            
            conn.close()
            
            return {
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'summary': {
                    'total_waste_kg': total_waste,
                    'total_waste_tonnes': round(total_waste / 1000, 2)
                },
                'by_type': by_type,
                'by_ward': by_ward
            }
        
        except Exception as e:
            return {'error': str(e)}
    
    @staticmethod
    def generate_csv_report(report_type: str, data: dict) -> str:
        """Generate CSV format report"""
        try:
            import csv
            from io import StringIO
            
            output = StringIO()
            writer = csv.writer(output)
            
            if report_type == 'complaints':
                writer.writerow(['Complaint Summary Report', data['period']['start'], 
                               'to', data['period']['end']])
                writer.writerow([])
                writer.writerow(['Metric', 'Value'])
                
                for key, value in data['summary'].items():
                    writer.writerow([key.replace('_', ' ').title(), value])
                
                writer.writerow([])
                writer.writerow(['By Type', 'Count'])
                for wtype, count in data['by_type'].items():
                    writer.writerow([wtype, count])
                
                writer.writerow([])
                writer.writerow(['By Ward', 'Count'])
                for ward, count in data['by_ward'].items():
                    writer.writerow([ward, count])
            
            elif report_type == 'waste':
                writer.writerow(['Waste Collection Report', data['period']['start'], 
                               'to', data['period']['end']])
                writer.writerow([])
                writer.writerow(['Metric', 'Value'])
                
                for key, value in data['summary'].items():
                    writer.writerow([key.replace('_', ' ').title(), value])
                
                writer.writerow([])
                writer.writerow(['By Type', 'Quantity (kg)', 'Collections'])
                for wtype, values in data['by_type'].items():
                    writer.writerow([wtype, values['quantity'], values['collections']])
                
                writer.writerow([])
                writer.writerow(['By Ward', 'Quantity (kg)', 'Collections'])
                for ward, values in data['by_ward'].items():
                    writer.writerow([ward, values['quantity'], values['collections']])
            
            return output.getvalue()
        
        except Exception as e:
            return f"Error generating CSV: {str(e)}"
    
    @staticmethod
    def generate_html_report(report_type: str, data: dict) -> str:
        """Generate HTML format report"""
        try:
            report_title = {
                'complaints': 'Complaint Summary Report',
                'waste': 'Waste Collection Report',
            }.get(report_type, f"{report_type.title()} Report")

            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>RCPI Waste AI - {report_title}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
                    .container {{ max-width: 900px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 8px; }}
                    h1 {{ color: #667eea; border-bottom: 3px solid #667eea; padding-bottom: 10px; }}
                    h2 {{ color: #764ba2; margin-top: 30px; }}
                    .summary {{ background-color: #f9f9f9; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                    .metric {{ display: inline-block; width: 200px; padding: 10px; margin: 5px; background-color: #e8e8f5; border-radius: 4px; }}
                    .metric-label {{ font-weight: bold; color: #667eea; }}
                    .metric-value {{ font-size: 24px; color: #333; }}
                    table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                    th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
                    th {{ background-color: #667eea; color: white; }}
                    tr:hover {{ background-color: #f5f5f5; }}
                    .footer {{ margin-top: 30px; color: #999; font-size: 12px; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>📊 {report_title}</h1>
                    <p>Period: {data['period']['start']} to {data['period']['end']}</p>
                    
                    <h2>Summary Metrics</h2>
                    <div class="summary">
            """
            
            for key, value in data.get('summary', {}).items():
                label = key.replace('_', ' ').title()
                html += f'<div class="metric"><div class="metric-label">{label}</div><div class="metric-value">{value}</div></div>'
            
            html += """
                    </div>
                    
                    <h2>Details by Category</h2>
            """
            
            if data.get('by_type'):
                html += """
                    <table>
                        <tr><th>Type</th><th>Quantity/Count</th></tr>
                """
                for key, value in data['by_type'].items():
                    if isinstance(value, dict):
                        html += f"<tr><td>{key}</td><td>{value.get('quantity', value.get('collections', 0))}</td></tr>"
                    else:
                        html += f"<tr><td>{key}</td><td>{value}</td></tr>"
                html += "</table>"
            
            html += """
                    <div class="footer">
                        <p>Generated: """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """</p>
                        <p>RCPI Waste Management System © 2024</p>
                    </div>
                </div>
            </body>
            </html>
            """
            
            return html
        
        except Exception as e:
            return f"Error generating HTML: {str(e)}"
    
    @staticmethod
    def generate_json_report(report_type: str, data: dict) -> str:
        """Generate JSON format report"""
        try:
            report = {
                'report_type': report_type,
                'generated_at': datetime.now().isoformat(),
                'data': data
            }
            return json.dumps(report, indent=2, default=str)
        except Exception as e:
            return f"Error generating JSON: {str(e)}"

class PDFReportBuilder:
    """Build PDF reports (requires reportlab library)"""
    
    @staticmethod
    def generate_pdf_report(report_type: str, data: dict) -> Optional[bytes]:
        """Generate PDF report"""
        try:
            # This requires: pip install reportlab
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
            from io import BytesIO
            
            # Create PDF in memory
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#667eea'),
                spaceAfter=30,
                alignment=1
            )
            
            story.append(Paragraph(f"RCPI Waste AI - {report_type.title()} Report", title_style))
            story.append(Spacer(1, 0.3*inch))
            
            # Period
            period_text = f"Period: {data['period']['start']} to {data['period']['end']}"
            story.append(Paragraph(period_text, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            # Summary table
            if data.get('summary'):
                summary_data = [['Metric', 'Value']]
                for key, value in data['summary'].items():
                    summary_data.append([key.replace('_', ' ').title(), str(value)])
                
                summary_table = Table(summary_data)
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(summary_table)
                story.append(Spacer(1, 0.3*inch))
            
            # Footer
            story.append(Spacer(1, 0.5*inch))
            footer_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            story.append(Paragraph(footer_text, styles['Normal']))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
        
        except ImportError:
            return None
        except Exception as e:
            return None

class ExcelReportBuilder:
    """Build Excel reports (requires openpyxl library)"""
    
    @staticmethod
    def generate_excel_report(report_type: str, data: dict) -> Optional[bytes]:
        """Generate Excel report"""
        try:
            # This requires: pip install openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Title
            ws['A1'] = f"RCPI Waste AI - {report_type.title()} Report"
            ws['A1'].font = Font(size=16, bold=True, color="667eea")
            ws.merge_cells('A1:D1')
            
            # Period
            ws['A3'] = f"Period: {data['period']['start']} to {data['period']['end']}"
            ws.merge_cells('A3:D3')
            
            # Summary
            row = 5
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            
            for key, value in data.get('summary', {}).items():
                row += 1
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
            
            # Set column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            
            # Convert to bytes
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        
        except ImportError:
            return None
        except Exception as e:
            return None
